import os
import re
import json
import logging
from datetime import datetime, timezone

import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
from openpyxl import load_workbook

NETWORK_ACTIVITY_COLUMN = 4
HOURS_COLUMN = 7
HEADER_ROW_COUNT = 2
MAX_HOURS = 100_000

JIRA_BASE_URL = "https://tt-rtx-26.atlassian.net"
PROJECT_KEY = "KAN"
EPIC_ISSUE_TYPE = "Epic"
EPIC_NAME_PATTERN = re.compile(r"^Epic (\d+)$")

SNAPSHOT_FILE = "snapshot.json"
EXCEL_FILE = "data.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
log = logging.getLogger(__name__)


class ExcelLoader:
    """Loads spreadsheet rows from a committed .xlsx file using openpyxl."""

    def __init__(self, filepath: str = EXCEL_FILE):
        self.filepath = filepath

    def load(self) -> list:
        """Returns rows as a 2D list of strings, skipping the two header rows."""
        wb = load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=HEADER_ROW_COUNT + 1, values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()
        return rows


class SnapshotManager:
    """Persists last-synced state to a JSON file for change detection across runs."""

    def __init__(self, filepath: str = SNAPSHOT_FILE):
        self.filepath = filepath

    def load(self) -> dict:
        """Returns saved activity states, or empty dict on first run."""
        if not os.path.exists(self.filepath):
            return {}
        with open(self.filepath, "r") as f:
            return json.load(f).get("activities", {})

    def save(self, activities: dict) -> None:
        with open(self.filepath, "w") as f:
            json.dump({"last_sync": datetime.now(timezone.utc).isoformat(), "activities": activities}, f, indent=2)
        log.info(f"Snapshot saved ({len(activities)} activities).")


class Validator:
    """Validates a spreadsheet row before it is synced to Jira."""

    def validate(self, activity: str, raw_hours: str) -> tuple[bool, str]:
        """Returns (True, '') if valid, or (False, reason) otherwise."""
        if not activity:
            return False, "missing activity name"
        if not raw_hours:
            return False, f"missing hours for '{activity}'"
        try:
            hours = float(raw_hours)
        except ValueError:
            return False, f"non-numeric hours '{raw_hours}' for '{activity}'"
        if hours < 0:
            return False, f"negative hours {hours} for '{activity}'"
        if hours > MAX_HOURS:
            return False, f"hours {hours} exceeds maximum {MAX_HOURS} for '{activity}'"
        return True, ""


class ActivityFinder:
    """Extracts a validated {network activity: hours} mapping from raw spreadsheet rows."""

    def __init__(self, data: list):
        self.data = data
        self._validator = Validator()

    def unique_activities(self) -> dict:
        """Returns an ordered dict of valid activities. First occurrence wins for duplicates."""
        activities = {}
        for line in self.data:
            if len(line) <= max(NETWORK_ACTIVITY_COLUMN, HOURS_COLUMN):
                continue

            name = line[NETWORK_ACTIVITY_COLUMN].strip()
            if name in activities:
                continue

            raw_hours = line[HOURS_COLUMN].strip()
            valid, reason = self._validator.validate(name, raw_hours)
            if not valid:
                log.warning(f"Skipping invalid row: {reason}.")
                continue

            activities[name] = float(raw_hours)
        return activities


class JiraClient:
    """Handles communication with the Jira REST API."""

    ACTUAL_HOURS_FIELD = "customfield_10104"
    NETWORK_ACTIVITY_FIELD = "customfield_10105"

    def __init__(self, base_url: str, email: str, api_token: str):
        self.base_url = base_url
        self.auth = HTTPBasicAuth(email, api_token)
        self.headers = {"Accept": "application/json", "Content-Type": "application/json"}

    def get_epics(self, project_key: str) -> list:
        """Returns all epics in the project with their custom field values."""
        url = f"{self.base_url}/rest/api/3/search/jql"
        jql = f"project = {project_key} AND issuetype = {EPIC_ISSUE_TYPE}"
        fields = ["summary", self.NETWORK_ACTIVITY_FIELD, self.ACTUAL_HOURS_FIELD]

        epics = []
        next_page_token = None
        while True:
            payload = {"jql": jql, "fields": fields, "maxResults": 100}
            if next_page_token:
                payload["nextPageToken"] = next_page_token

            response = requests.post(url, json=payload, auth=self.auth, headers=self.headers)
            if response.status_code != 200:
                log.error(f"Failed to fetch epics: {response.status_code} {response.text}")
                break

            body = response.json()
            for issue in body.get("issues", []):
                f = issue.get("fields", {})
                epics.append({
                    "key": issue["key"],
                    "summary": f.get("summary"),
                    "network_activity": f.get(self.NETWORK_ACTIVITY_FIELD),
                    "hours": f.get(self.ACTUAL_HOURS_FIELD),
                })

            if body.get("isLast", True):
                break
            next_page_token = body.get("nextPageToken")
            if not next_page_token:
                break

        return epics

    def create_epic(self, project_key: str, summary: str) -> str | None:
        """Creates a new epic and returns its key, or None on failure."""
        url = f"{self.base_url}/rest/api/3/issue"
        payload = {"fields": {
            "project": {"key": project_key},
            "issuetype": {"name": EPIC_ISSUE_TYPE},
            "summary": summary,
        }}
        response = requests.post(url, json=payload, auth=self.auth, headers=self.headers)
        if response.status_code == 201:
            key = response.json()["key"]
            log.info(f"Created epic {key} ({summary}).")
            return key
        log.error(f"Failed to create epic '{summary}': {response.status_code} {response.text}")
        return None

    def update_epic_fields(self, epic_key: str, hours: float = None, network_activity: str = None) -> bool:
        """Updates Actual Hours and/or Network Activity on an epic."""
        fields = {}
        if hours is not None:
            fields[self.ACTUAL_HOURS_FIELD] = float(hours)
        if network_activity is not None:
            fields[self.NETWORK_ACTIVITY_FIELD] = network_activity

        if not fields:
            return True

        response = requests.put(
            f"{self.base_url}/rest/api/3/issue/{epic_key}",
            json={"fields": fields}, auth=self.auth, headers=self.headers,
        )
        if response.status_code == 204:
            parts = ([f"Network Activity={network_activity}"] if network_activity else []) + \
                    ([f"Actual Hours={hours}"] if hours is not None else [])
            log.info(f"Updated {epic_key}: {', '.join(parts)}.")
            return True
        log.error(f"Failed to update {epic_key}: {response.status_code} {response.text}")
        return False


def next_epic_number(epics: list) -> int:
    highest = 0
    for epic in epics:
        m = EPIC_NAME_PATTERN.match((epic.get("summary") or "").strip())
        if m:
            highest = max(highest, int(m.group(1)))
    return highest + 1


def hours_match(a, b) -> bool:
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def main():
    load_dotenv()

    loader = ExcelLoader()
    data = loader.load()
    finder = ActivityFinder(data)
    activities = finder.unique_activities()

    if not activities:
        log.info("No valid network activities found in the spreadsheet.")
        return

    jira = JiraClient(
        base_url=JIRA_BASE_URL,
        email=os.getenv("EMAIL"),
        api_token=os.getenv("API_KEY"),
    )

    epics = jira.get_epics(PROJECT_KEY)
    epics_by_activity = {e["network_activity"]: e for e in epics if e["network_activity"]}
    epic_counter = next_epic_number(epics)

    snapshot_mgr = SnapshotManager()
    snapshot = snapshot_mgr.load()
    new_snapshot = dict(snapshot)

    conflicts = []

    log.info(f"Processing {len(activities)} activities against {len(epics)} existing epics.")

    for activity, sheet_hours in activities.items():
        existing = epics_by_activity.get(activity)
        snap = snapshot.get(activity)
        snap_hours = snap.get("hours") if snap else None

        # ── New activity: create epic and populate both fields ──────────────
        if not existing:
            summary = f"Epic {epic_counter}"
            epic_counter += 1
            new_key = jira.create_epic(PROJECT_KEY, summary)
            if new_key:
                jira.update_epic_fields(new_key, hours=sheet_hours, network_activity=activity)
                new_snapshot[activity] = {"hours": sheet_hours, "jira_key": new_key}
            continue

        jira_key = existing["key"]
        jira_hours = existing.get("hours")

        sheet_changed = snap_hours is None or not hours_match(snap_hours, sheet_hours)
        # Jira-side change is only meaningful when we have a prior snapshot to compare against.
        jira_changed = snap_hours is not None and not hours_match(snap_hours, jira_hours)

        # ── True conflict: both sides diverged to different values ──────────
        if sheet_changed and jira_changed and not hours_match(sheet_hours, jira_hours):
            conflicts.append({
                "epic": jira_key, "activity": activity,
                "sheet": sheet_hours, "jira": jira_hours, "last_sync": snap_hours,
            })
            log.warning(
                f"CONFLICT {jira_key} ('{activity}'): "
                f"sheet={sheet_hours}, jira={jira_hours}, last_sync={snap_hours}. "
                f"Skipping — resolve manually then re-run."
            )
            continue

        # ── Jira-side edit: Jira diverged while sheet stayed the same ───────
        if jira_changed and not sheet_changed:
            log.warning(
                f"Jira-side edit on {jira_key} ('{activity}'): "
                f"jira={jira_hours} (was {snap_hours}). "
                f"Sheet unchanged at {sheet_hours} — restoring sheet value."
            )
            if jira.update_epic_fields(jira_key, hours=sheet_hours):
                new_snapshot[activity] = {"hours": sheet_hours, "jira_key": jira_key}
            continue

        # ── Sheet changed: normal sync ───────────────────────────────────────
        if sheet_changed:
            if jira.update_epic_fields(jira_key, hours=sheet_hours):
                new_snapshot[activity] = {"hours": sheet_hours, "jira_key": jira_key}
            continue

        # ── Nothing changed ──────────────────────────────────────────────────
        log.info(f"{jira_key} up to date for '{activity}' ({sheet_hours} hours).")
        new_snapshot[activity] = {"hours": sheet_hours, "jira_key": jira_key}

    snapshot_mgr.save(new_snapshot)

    if conflicts:
        log.warning(f"\n{len(conflicts)} unresolved conflict(s):")
        for c in conflicts:
            log.warning(f"  {c['epic']} ('{c['activity']}'): sheet={c['sheet']}, jira={c['jira']}, last_sync={c['last_sync']}")

    log.info("Done.")


if __name__ == "__main__":
    main()
